import openpyxl
import urllib
wb = openpyxl.load_workbook('data_download.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
link = ''
category = 0
prev_category = -1
data_labels = {}
idx = 0
out_file = open('data/data_label.txt','w')
out_file.write('FileName ClassId\n')
for i in range(2, 13237): #13236
    if sheet.cell(row=i, column=1).value == '============' or sheet.cell(row=i,column=2).value == '------------':
        continue
    else:
        hyperlink = sheet.cell(row=i, column=2).value
        link = sheet.cell(row=i, column=2).value[12:len(hyperlink)-9]
        category = sheet.cell(row=i, column=5).value
        if prev_category == category:
            idx+=1
        else:
            idx = 1
            prev_category = category
        #print(i, sheet.cell(row=i, column=2).value)

        f = 'data/' + str(category) + '_' + str(idx) + '.mov'
        data_labels.setdefault(category, []).append(f)
        out_file.write(f+' '+str(category)+'\n')
        urllib.urlretrieve(link, filename=f)



